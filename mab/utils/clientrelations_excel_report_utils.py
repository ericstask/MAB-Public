import os
import re
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta


def get_previous_byday(dayname, start_date=None):
    weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    if start_date is None:
        start_date = datetime.today()
    day_num = start_date.weekday()
    day_num_target = weekdays.index(dayname)
    days_ago = (7 + day_num - day_num_target) % 7  
    if days_ago == 0:
        days_ago = 7
    target_date = start_date - timedelta(days=days_ago)
    return target_date


def read_file_data(filename, skip_header=False):
    try:
        with open(filename) as file:
            print(f"{filename} file found.")
            if skip_header:
                lines = file.readlines()[1:] # skip header
            else:
                lines = file.readlines()
            return [line.strip().split('\t') for line in lines]
    except FileNotFoundError:
        print(f"{filename} file not found. Press ENTER to exit.")
        input()
        sys.exit()


def create_excel_sheet(dict_data, create_excel_header, format_excel_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Active"

    create_excel_header(ws)
    format_excel_data(ws, dict_data)

    return wb


# REPORT SPECIFIC FUNCTIONS ##################################################################################
        
def bank_sif_report_create_excel_header(ws):
    # Header formatting
    header_style = Font(size=11, bold=True, name="Times New Roman")
    standard_font = Font(size=11, name="Times New Roman")
    fill_color = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
    border = Border(top=Side(style='thick'))
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws["E1"] = "Settlement Tracker"
    ws["E1"].font = header_style
    ws["E1"].alignment = Alignment(horizontal='center')

    ws["A2"] = "Week Of:"
    ws["A2"].font = standard_font
    ws["A2"].alignment = Alignment(horizontal='center')

    last_friday = get_previous_byday("Friday").strftime('%m-%d-%y')
    last_thursday = get_previous_byday("Thursday").strftime('%m-%d-%y')
    ws["B2"] = last_friday + " to " + last_thursday
    ws["B2"].font = standard_font
    ws["B2"].alignment = Alignment(horizontal='center')

    # Header content
    headers = ["Internal Account #", "Account #", "Customer's Name", "Balance at time of Negotiation",
               "Settlement Amount", "Settlement Percentage", "Settlement Approval Authority",
               "Settlement Paid Date", "Date Letter Sent to Customer", "Stream", "Placement Date"]

    # Set column and row widths
    widths = [17, 19, 24, 12, 20, 15, 12, 18, 21, 16, 23]
    # for col, width in zip(ws.iter_cols(min_col=1, max_col=len(headers)), widths):
    for col, width in zip(range(1, len(headers) + 1), widths):
        # col[0].width = width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Write header content
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=idx, value=header)
        cell.font = header_style
        cell.alignment = alignment
        cell.fill = fill_color
        cell.border = border


def bank_sif_report_format_excel_data(ws, dict_data):
    row = 4
    for data in dict_data:
        if data["CLIENT#"] == 'BANCACS3':
            continue
        
        ws["A" + str(row)] = data["ACCOUNT#"] 
        ws["B" + str(row)] = data["C.D.#"]
        ws["C" + str(row)] = data["NAME 1"]
        ws["D" + str(row)] = float(data["PN-ASSIGN"])
        ws["E" + str(row)] = float(data["PRIN COL"])
        ws["H" + str(row)] = data["LST PAY"]
        ws["I" + str(row)] = data["LST LTR/NTC DT"]
        ws["J" + str(row)] = data["CLIENT#"]
        ws["K" + str(row)] = data["DT-ASSIGN"]

        ws["F" + str(row)] = float(data["PRIN COL"]) / float(data["PN-ASSIGN"])
        ws["F" + str(row)].number_format = '0.00%'

        if data["CLIENT#"] in ["BANCACS1", "BANCC1", "BANED1"]:
            if float(data["PN-ASSIGN"]) < 2000.0:
                ws["G" + str(row)] = 0.6
                ws["G" + str(row)].number_format = '0%'
            if float(data["PN-ASSIGN"]) >= 2000.0 and float(data["PN-ASSIGN"]) < 8000.0:
                ws["G" + str(row)] = 0.5
                ws["G" + str(row)].number_format = '0%'
            if float(data["PN-ASSIGN"]) >= 8000.0 and float(data["PN-ASSIGN"]) < 25000.0:
                ws["G" + str(row)] = 0.45
                ws["G" + str(row)].number_format = '0%'
            if float(data["PN-ASSIGN"]) >= 25000.0:
                ws["G" + str(row)] = 0.40
                ws["G" + str(row)].number_format = '0%'
        if data["CLIENT#"] in ["BANCACS2", "BANCC2", "BANED2"]:
            if float(data["PN-ASSIGN"]) < 8000.0:
                ws["G" + str(row)] = 0.40
                ws["G" + str(row)].number_format = '0%'
            if float(data["PN-ASSIGN"]) >= 8000.0:
                ws["G" + str(row)] = 0.30
                ws["G" + str(row)].number_format = '0%'
                
        row += 1


def create_bank_sif_report(file_directory):
    last_friday = get_previous_byday("Friday").strftime('%m-%d-%y')
    last_thursday = get_previous_byday("Thursday").strftime('%m-%d-%y')
    filename = next((file for file in os.listdir(file_directory) if "BANSIF" in file), None)
    file_path = os.path.join(file_directory, filename)

    if filename:
        file_data = read_file_data(file_path)
        keys = file_data.pop(0)
        keys[9] = keys[9].rstrip()  # Remove the newline character
        dict_data = [dict(zip(keys, data)) for data in file_data]

        for line in dict_data:
            print(line)

        wb = create_excel_sheet(dict_data, bank_sif_report_create_excel_header, bank_sif_report_format_excel_data)
        save_file_path = os.path.join(file_directory, f"Settlement Tracker {last_friday} to {last_thursday}.xlsx")
        wb.save(save_file_path)
        print("Excel spreadsheet created successfully.")
    else:
        print("BANSIF file not found.")


def create_revo_weekly_invoice_worksheet(portfolio, data, today, file_directory):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    ### --- Formatting for Heading --- ###
    BoldCalibri14 = Font(size=14, bold=True, name="Calibri")
    BlueCalibri = Font(size=11, bold=True, name="Calibri", color="0070C0")
    RedCalibri = Font(size=11, bold=True, name="Calibri", color="FF0000")
    StandardCalibri = Font(size=11, name="Calibri", color="000000")

    worksheet.column_dimensions["A"].width = 17
    worksheet.column_dimensions["B"].width = 17
    worksheet.column_dimensions["C"].width = 17
    worksheet.column_dimensions["D"].width = 17
    worksheet.column_dimensions["E"].width = 17
    worksheet.column_dimensions["F"].width = 17
    worksheet.column_dimensions["G"].width = 17
    worksheet.column_dimensions["H"].width = 17
    worksheet.column_dimensions["I"].width = 17
    worksheet.column_dimensions["J"].width = 17
    worksheet.column_dimensions["K"].width = 17

    worksheet["A1"] = "PAYMENT INVOICE"
    worksheet["A1"].font = BoldCalibri14
    worksheet.merge_cells("A1:B1")

    worksheet["A2"] = "Client:" ####
    worksheet["A2"].font = BlueCalibri
    if portfolio == "98531":
        worksheet["B2"] = "Revo"
    elif portfolio == "98532":
        worksheet["B2"] = "Ervo"
    elif portfolio == "98533":
        worksheet["B2"] = "Vreo"
    elif portfolio == "98534":
        worksheet["B2"] = "Orev"
    elif portfolio == "98535":
        worksheet["B2"] = "Over"
    elif portfolio == "98536":
        worksheet["B2"] = "Voer"
    elif portfolio == "98537":
        worksheet["B2"] = "Erov"
    elif portfolio == "98538":
        worksheet["B2"] = "Rove"
    elif portfolio == "98539":
        worksheet["B2"] = "Vroe"
    elif portfolio == "98540":
        worksheet["B2"] = "Oerv"
    else: 
        print("\t Portfolio code not recognized.")
    

    worksheet["A3"] = "Reporting Period:"
    worksheet["A3"].font = BlueCalibri
    lastMonday = get_previous_byday("Monday")
    lastFriday = get_previous_byday("Friday")
    worksheet["B3"] = lastMonday.strftime('%m/%d/%y') + " - " + lastFriday.strftime('%m/%d/%y')

    worksheet["E2"] = "Portfolio #"
    worksheet["E2"].font = BlueCalibri
    worksheet["F2"] = portfolio

    worksheet["E3"] = "Invoice Date"
    worksheet["E3"].font = BlueCalibri
    worksheet["F3"] = today.strftime('%m/%d/%Y')

    worksheet["H2"] = "Page:"
    worksheet["H2"].font = BlueCalibri
    worksheet["I2"] = "1 of 1"

    worksheet["H3"] = "Invoice #:"
    worksheet["H3"].font = BlueCalibri
    worksheet["I3"] = today.strftime('%m-%d-%y') + "-" + portfolio

    worksheet["A4"] = portfolio

    ### --- Formatting for Data --- ###

    Bold = Font(bold=True)
    ThinBorder = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
    CenterAlignment = Alignment(horizontal="center")

    worksheet["A5"] = "Customer's Acct #"
    worksheet["A5"].border = ThinBorder
    worksheet["A5"].alignment = CenterAlignment
    worksheet["A5"].font = Bold

    worksheet["B5"] = "Name"
    worksheet["B5"].border = ThinBorder
    worksheet["B5"].alignment = CenterAlignment
    worksheet["B5"].font = Bold

    worksheet["C5"] = "Address (optional)"
    worksheet["C5"].border = ThinBorder
    worksheet["C5"].alignment = CenterAlignment
    worksheet["C5"].font = RedCalibri

    worksheet["D5"] = "Pay Date"
    worksheet["D5"].border = ThinBorder
    worksheet["D5"].alignment = CenterAlignment
    worksheet["D5"].font = Bold

    worksheet["E5"] = "Paid Us"
    worksheet["E5"].border = ThinBorder
    worksheet["E5"].alignment = CenterAlignment
    worksheet["E5"].font = Bold

    worksheet["F5"] = "Paid You"
    worksheet["F5"].border = ThinBorder
    worksheet["F5"].alignment = CenterAlignment
    worksheet["F5"].font = Bold

    worksheet["G5"] = "Comission"
    worksheet["G5"].border = ThinBorder
    worksheet["G5"].alignment = CenterAlignment
    worksheet["G5"].font = Bold

    worksheet["H5"] = "Current Balance"
    worksheet["H5"].border = ThinBorder
    worksheet["H5"].alignment = CenterAlignment
    worksheet["H5"].font = Bold

    worksheet["I5"] = "Remarks (optional)"
    worksheet["I5"].border = ThinBorder
    worksheet["I5"].alignment = CenterAlignment
    worksheet["I5"].font = RedCalibri

    worksheet["J5"] = "Types of Payment (optional)"
    worksheet["J5"].border = ThinBorder
    worksheet["J5"].alignment = CenterAlignment
    worksheet["J5"].font = RedCalibri

    worksheet["K5"] = "Tax Amount"
    worksheet["K5"].border = ThinBorder
    worksheet["K5"].alignment = CenterAlignment
    worksheet["K5"].font = Bold

    # Add data
    if data:
        row = 6
        for datum in data:
            for i, val in enumerate(datum, 1):
                worksheet[get_column_letter(i) + str(row)] = val
            row += 1

        row += 1
        worksheet["D" + str(row)] = "TOTALS"
        worksheet["D" + str(row)].font = Bold

        totals = [sum(float(datum[i]) for datum in data) for i in range(4, 7)]
        for i, total in enumerate(totals, 4):
            worksheet[get_column_letter(i) + str(row)] = f"{total:.2f}"
            worksheet[get_column_letter(i) + str(row)].font = Bold

    else:
        worksheet["A6"] = "NOTHING TO REPORT"
        print("No data found in", portfolio)


    workbook.save(os.path.join(file_directory, f"Agency Invoice Merc{today.strftime('%m-%d-%y')} {portfolio}.xlsx"))


def create_revo_weekly_invoices(file_directory):
    today = datetime.today()
    file_names = ["remit_98531", "remit_98532", "remit_98533", "remit_98534", "remit_98535", "remit_98536", "remit_98537", "remit_98538", "remit_98539", "remit_98540"]
    print("Finding CUBS Files...")
    files = [file for file_name in file_names for file in os.listdir(file_directory) if file.startswith(file_name)]
    file_data = {file: read_file_data(os.path.join(file_directory, file), True) for file in files}
    print(file_data)

    for file_name, data in file_data.items():
        portfolio = re.search(r'_(.*?)_', file_name).group(1)
        create_revo_weekly_invoice_worksheet(portfolio, data, today, file_directory)


def create_quarterly_worksheet(data, output_folder):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Call Logs"
    bold10ArialFont = Font(size=10, bold=True, name="Arial")
    greenColor = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    for c in range(1, 12):
        sheet[get_column_letter(c) + '1'].font = Font(size=10, name="Arial")
        sheet[get_column_letter(c) + '1'].fill = greenColor
        sheet.column_dimensions[get_column_letter(c)].width = 20

    sheet['A1'] = "Vendor"
    sheet['B1'] = "Account_Number"
    sheet['C1'] = "Customer_Address"
    sheet['D1'] = "Customer_Zipcode"
    sheet['E1'] = "Number_Dialed"
    sheet['F1'] = "Call_Date_Time"
    sheet['G1'] = "Call_Time_Timezone"
    sheet['H1'] = "Action"
    sheet['I1'] = "Result_Description"
    sheet['J1'] = "Customer_Timezone based on Zip Code"
    sheet['K1'] = "Customer_Timezone based on Area Code"

    for row, item in enumerate(data, start=2):
        sheet['A' + str(row)] = "Western"
        sheet['B' + str(row)] = item[8]
        sheet['C' + str(row)] = item[10]
        sheet['D' + str(row)] = item[12]
        sheet['E' + str(row)] = item[13]
        sheet['F' + str(row)] = f"{item[2]} {item[14]}"
        sheet['G' + str(row)] = "Eastern"
        sheet['H' + str(row)] = item[18]
        sheet['I' + str(row)] = item[21]
        sheet['J' + str(row)] = item[23]
        sheet['K' + str(row)] = item[24]

    file_path = os.path.join(output_folder, f"AClient_Quarterly_Call_Log_MAB.xlsx")
    workbook.save(file_path)


def create_monthly_worksheet(data, output_folder):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Call Logs"
    bold10ArialFont = Font(size=10, bold=True, name="Arial")
    grayColor = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

    for c in range(1, 13):
        sheet[get_column_letter(c) + '1'].font = Font(size=10, bold=True, name="Arial")
        sheet[get_column_letter(c) + '1'].fill = grayColor
        sheet.column_dimensions[get_column_letter(c)].width = 20

    sheet['A1'] = "Vendor"
    sheet['B1'] = "Client"
    sheet['C1'] = "Account_Number"
    sheet['D1'] = "Customer_Address"
    sheet['E1'] = "Customer_Zipcode"
    sheet['F1'] = "Number_Dialed"
    sheet['G1'] = "Call_Date_Time"
    sheet['H1'] = "Call_Time_Timezone"
    sheet['I1'] = "Action"
    sheet['J1'] = "Result_Description  Inbound , outbound, vmail"
    sheet['K1'] = "Customer_Timezone based on Zip Code"
    sheet['L1'] = "Customer_Timezone based on Area Code"

    for x, row_data in enumerate(data):
        sheet['A' + str(x+2)] = "Western"
        
        if row_data[4] == "aclient01":
            sheet['B' + str(x+2)] = "Retail"
        elif row_data[4] == "aclient02":
            sheet['B' + str(x+2)] = "Lease"
        else:
            sheet['B' + str(x+2)] = "ERROR"
        
        sheet['C' + str(x+2)] = row_data[8]
        sheet['D' + str(x+2)] = row_data[10]
        sheet['E' + str(x+2)] = row_data[12]
        sheet['F' + str(x+2)] = row_data[13]
        sheet['G' + str(x+2)] = row_data[2] + " " + row_data[14]
        sheet['H' + str(x+2)] = "Eastern"
        sheet['I' + str(x+2)] = row_data[18]
        sheet['J' + str(x+2)] = row_data[21]
        sheet['K' + str(x+2)] = row_data[23]
        sheet['L' + str(x+2)] = row_data[24]

    file_path = os.path.join(output_folder, f"AClient_Monthly_Call_Log_MAB.xlsx")
    workbook.save(file_path)


def create_aclient_call_logs(report_type, file_directory):
    if report_type == "Quarterly": 
        file_list = [file for file in os.listdir(file_directory) if file.startswith('AClient_call_') and '.' not in file.split('.')[0]]
        data = []
        for filename in file_list:
            data += read_file_data(os.path.join(file_directory, filename), True)
        create_quarterly_worksheet(data, file_directory)
    elif report_type == "Monthly": 
        file_list = [file for file in os.listdir(file_directory) if file.startswith('AClient_call_') and '.' not in file.split('.')[0]]
        data = []
        for filename in file_list:
            data += read_file_data(os.path.join(file_directory, filename), True)
        create_monthly_worksheet(data, file_directory)
        

def create_sectional_call_logs(file_directory):
    # Find PCI file
    pci_files = [file for file in os.listdir(file_directory) if file.startswith("Sectional_CALL")]
    # if len(pci_files) != 1:
    #     print("There should be exactly one 'Sectional_CALL' file in the folder.")
    #     sys.exit(1)

    print(pci_files)
    # Extract data from the PCI file
    CallReportData = []
    with open(os.path.join(file_directory, pci_files[0])) as file:
        lines = file.readlines()
        header = lines.pop(0).strip().split('\t')
        for line in lines:
            datum = line.strip().split('\t')
            if int(datum[16]) < 60:
                continue  # Skip lines with a duration less than a minute
            CallReportData.append(dict(zip(header, datum)))

    # Create spreadsheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Create headings for excel sheet
    headings = [
        "AcctNbr", "ClientID", "CallDate", "CallStartTime", "CallEndTime",
        "Duration", "PhoneNbr", "PhoneType", "CallType", "State", "Zip"
    ]
    for col, heading in enumerate(headings, start=1):
        cell = ws.cell(row=1, column=col, value=heading)
        cell.font = Font(bold=True)

    # Fill in data
    for row, data in enumerate(CallReportData, start=2):
        ws.cell(row=row, column=1, value=data["Account No"])
        ws.cell(row=row, column=2, value=data["Client"])
        ws.cell(row=row, column=3, value=datetime.strptime(data["Call Date"], '%m/%d/%Y').strftime('%Y/%m/%d'))
        ws.cell(row=row, column=4, value=data["Call Start Time"])
        ws.cell(row=row, column=5, value=data["Call End Time"])
        ws.cell(row=row, column=6, value=int(data["Duration"]))
        ws.cell(row=row, column=7, value=data["Phone"])
        ws.cell(row=row, column=8, value=data["Phone Type"])
        ws.cell(row=row, column=9, value="OUT" if data["IB/OB"] == "OUTBOUND" else data["IB/OB"])
        ws.cell(row=row, column=10, value=data["Account State"])
        ws.cell(row=row, column=11, value=data["Account Zip"].split('-', 1)[0])

    # Save the file with a name indicating the last month
    today = datetime.today()
    last_month = (today.replace(day=1) - timedelta(days=1)).strftime('%B %Y')
    wb.save(os.path.join(file_directory, f"Sectional_CallLog_{last_month}.xlsx"))
