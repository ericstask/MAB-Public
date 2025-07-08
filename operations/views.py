import pandas as pd
import json
import requests
import os
from weasyprint import HTML
from datetime import datetime, timedelta, timezone
from dateutil.relativedelta import relativedelta
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.template.loader import render_to_string
from urllib.parse import urlencode
from .models import PayForPerformance, PayForPerformanceEmployee, AgentProductivity
from .forms import EmployeeForm, AgentProductivityForm
from .utils import create_pfp_spreadsheet, create_pfp_dataframe, pay_for_performance_query_transactions, agent_productivity_process_data

from io import BytesIO
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle

# Create your views here.
@login_required
@permission_required('operations.access_operations', raise_exception=True)
def reports(request):
    return render(request, 'operations/reports.html')


@login_required
@permission_required('operations.access_operations', raise_exception=True)
def tools(request):
    return render(request, 'operations/tools.html')


@login_required
@permission_required('operations.access_operations', raise_exception=True)
def pay_for_performance(request):
    # Upload file from paychex
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']

        # Make sure the file is an excel spreadsheet
        if not uploaded_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'The uploaded file was not an excel spreadsheet')
        else:
            required_columns = ['Employee ID', 'Employee Last Name and Suffix', 'Employee First Name', 'Most Recent Hire Date', 'Supervisor', 'Pay Rate 1',]
            df = pd.read_excel(uploaded_file) 
            missing_columns = [col for col in required_columns if col not in df.columns]

            # Make sure the spreadsheet has the required columns
            if missing_columns:
                messages.error(request, f'The uploaded spreadsheet is missing required columns: {', '.join(missing_columns)}')
            else:
                # Iterate through because a bulk delete will not trigger overridden delete function
                for instance in PayForPerformance.objects.all():
                    instance.delete()
                #PayForPerformance.objects.all().delete()

                file_instance = PayForPerformance(uploaded_file=uploaded_file, file_title=uploaded_file.name)
                file_instance.save()

                messages.success(request, 'File uploaded successfully.')

    context = {}

    file_instance = PayForPerformance.objects.last()

    # Add file information to context if their is one on file
    if file_instance:
        context["file_name"] = file_instance.file_title
        context["uploaded_date"] = file_instance.uploaded_at

    # Calucalte the start and end dates for the report
    endDate = datetime.today() - relativedelta(months=1)
    year = endDate.year
    month = endDate.month
    end_month = endDate.strftime('%Y-%m')

    # Override the year and month with values from dateInput in form if it exists in request
    if request.method == 'POST' and file_instance:
        year, month = request.POST.get('dateInput', end_month).split('-')

    end_date = datetime(int(year), int(month), 1) + timedelta(days=32)
    end_date = end_date.replace(day=1) - timedelta(days=1)

    start_date = datetime(int(year), int(month), 1) - relativedelta(months=5)

    # Convert datetime objects to strings to use in query
    startDate = start_date.strftime('%Y-%m-%d')
    endDate = end_date.strftime('%Y-%m-%d')

    # Create excel file
    if request.method == 'POST' and request.POST.get('create_pfp_excel_file'):
        file_path = file_instance.uploaded_file.path

        df = pd.read_excel(file_path)

        combined_df, transaction_data = pay_for_performance_query_transactions(df, startDate, endDate)

        final_data = create_pfp_dataframe(combined_df, transaction_data)

        pfp_spreadsheet_buffer = create_pfp_spreadsheet(final_data)

        response = HttpResponse(pfp_spreadsheet_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="pfp.xlsx"'

        return response

    context["selected_date"] = endDate.rsplit('-', 1)[0]    

    return render(request, 'operations/pay_for_performance.html', context)


@login_required
@permission_required('operations.access_operations', raise_exception=True)
def pay_for_performance_get_table_data(request):
    file_instance = PayForPerformance.objects.last()

    if not file_instance:
        return JsonResponse({}, content_type="application/json", status=204)
        return None
    
    # Calculate the required start and end months for report
    endDate = datetime.today() - relativedelta(months=1)
    year = endDate.year
    month = endDate.month
    end_month = endDate.strftime('%Y-%m')

    if request.method == 'POST' and file_instance:
        date_post_data = json.loads(request.body)
        dateInput = date_post_data['dateInput']

        if dateInput:
            year, month = str(dateInput).split('-')
        else:
            year, month = end_month.split('-')

    end_date = datetime(int(year), int(month), 1) + timedelta(days=32)
    end_date = end_date.replace(day=1) - timedelta(days=1)

    start_date = datetime(int(year), int(month), 1) - relativedelta(months=5)

    startDate = start_date.strftime('%Y-%m-%d')
    endDate = end_date.strftime('%Y-%m-%d')

    file_path = file_instance.uploaded_file.path

    df = pd.read_excel(file_path)

    combined_df, transaction_data = pay_for_performance_query_transactions(df, startDate, endDate)

    final_data = create_pfp_dataframe(combined_df, transaction_data)
    
    final_data_dict = final_data.to_dict(orient='records')

    return JsonResponse(final_data_dict, safe=False)


@login_required
@permission_required('operations.access_operations', raise_exception=True)
def pay_for_performance_add_remove(request):
    # Retrieve added and removed employees
    added_employees = PayForPerformanceEmployee.objects.filter(status='added', is_removed=False).order_by('first_name')
    removed_employees = PayForPerformanceEmployee.objects.filter(status='removed', is_removed=False).order_by('first_name')

    if request.method == 'POST':
        form = EmployeeForm(request.POST)

        if form.is_valid():
            action = form.cleaned_data['action']

            if action == 'add':
                employee = PayForPerformanceEmployee(
                    id=form.cleaned_data['id'],
                    last_name_and_suffix=form.cleaned_data['last_name_and_suffix'],
                    first_name=form.cleaned_data['first_name'],
                    most_recent_hire_date=form.cleaned_data['most_recent_hire_date'],
                    supervisor=form.cleaned_data['supervisor'],
                    pay_rate=form.cleaned_data['pay_rate'],
                    status='added',
                    is_removed=False,
                )

                employee.save()

                redirect('operations:pay_for_performance_add_remove')

            if action == 'remove':
                employee = PayForPerformanceEmployee(
                    id=form.cleaned_data['id'],
                    last_name_and_suffix=form.cleaned_data['last_name_and_suffix'],
                    first_name=form.cleaned_data['first_name'],
                    status='removed',
                    is_removed=False,
                )

                employee.save()

                redirect('operations:pay_for_performance_add_remove')
    
    form = EmployeeForm()

    context = {
        'form': form,
        'added_employees': added_employees,
        'removed_employees': removed_employees,
    }

    return render(request, 'operations/pay_for_performance_add_remove.html', context)


@login_required
@permission_required('operations.access_operations', raise_exception=True)
def pay_for_performance_delete(request, employee_id):
    employee = get_object_or_404(PayForPerformanceEmployee, id=employee_id)
    employee.is_removed = True
    employee.save()
    
    return redirect('operations:pay_for_performance_add_remove')


@login_required
@permission_required('operations.access_operations', raise_exception=True)
def agent_productivity(request):
    if request.method == 'POST':
        form = AgentProductivityForm(request.POST, request.FILES)

        if form.is_valid():
            
            productivity_file = form.cleaned_data['productivity_file']
            calls_file = form.cleaned_data['calls_file']

            if not productivity_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'The uploaded Agent Productivity Log is not an excel spreadsheet')
            if not calls_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'The uploaded Agent Call Log is not an excel spreadsheet')
            if productivity_file.name.endswith(('.xlsx', '.xls')) and calls_file.name.endswith(('.xlsx', '.xls')):
                productivity_required_columns = ['Agent', 'Total Calls', 'Inbound Calls', 'Transfers', 'Mobile Comply', 'PD', 'Outbound Manual Calls', 'In Call (Min)', 
                                                 'In Call (%)', 'Ready', 'Ready (%)', 'Wrapup', 'Wrapup (%)', 'Not Ready', 'Not Ready (%)', 'Total Login Time', 'Total RPC', 
                                                 'Total WPC', 'Non-Contacts',]
                productivity_df = pd.read_excel(productivity_file, skiprows=2) 
                productivity_missing_columns = [col for col in productivity_required_columns if col not in productivity_df.columns]

                calls_required_columns = ['Call Date & Time', 'Client', 'Agent', 'Record Key', 'Phone', 'Agent Duration', 'Call Duration', 'Result Code', 'Disposition', 
                                          'Result Note', 'Call Type', 'ClientReference01', 'ClientReference02',]
                
                calls_df = pd.read_excel(calls_file, skiprows=2) 
                calls_missing_columns = [col for col in calls_required_columns if col not in calls_df.columns]

                # Make sure the spreadsheet has the required columns
                if productivity_missing_columns:
                    messages.error(request, f'The uploaded Agent Productivity Log is missing required columns: {', '.join(productivity_missing_columns)}')
                if calls_missing_columns:
                    messages.error(request, f'The uploaded Agent Call Log is missing required columns: {', '.join(calls_missing_columns)}')
                if not productivity_missing_columns and not calls_missing_columns:

                    for instance in AgentProductivity.objects.all():
                        instance.delete()

                    file_instance = AgentProductivity(productivity_uploaded_file=productivity_file, productvity_file_title=productivity_file.name,
                                                    calls_uploaded_file=calls_file, calls_file_title=calls_file.name)

                    file_instance.save()

                    messages.success(request, f'Agent Productivity Log and Agent Call Log uploaded successfully')

                    redirect('operations:agent_productivity') 

    context = {}   
    
    files_instance = AgentProductivity.objects.last()
    if files_instance:
        context['last_upload'] = files_instance.uploaded_at

    form = AgentProductivityForm()
    context['form'] = form

    return render(request, 'operations/agent_productivity.html', context)


@login_required
@permission_required('operations.access_operations', raise_exception=True)
def agent_productivity_get_table_data(request):
    files_instance = AgentProductivity.objects.last()

    if not files_instance:
        return JsonResponse({}, content_type="application/json", status=204)

    merged_with_calls = agent_productivity_process_data(files_instance)

    final_data_dict = merged_with_calls.to_dict(orient='records')

    return JsonResponse(final_data_dict, safe=False)


@login_required
@permission_required('operations.access_operations', raise_exception=True)
def agent_productivity_create_excel_file(request):
    files_instance = AgentProductivity.objects.last()

    if not files_instance:
        return JsonResponse({}, content_type="application/json", status=204)

    merged_with_calls = agent_productivity_process_data(files_instance)

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        merged_with_calls.to_excel(writer, index=False)

    workbook = load_workbook(buffer)
    sheet = workbook.active

    # Define custom styles
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="9CAFCC", end_color="9CAFCC", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Change column widths to fit content
    for idx, col in enumerate(sheet.columns, 1):
        sheet.column_dimensions[get_column_letter(idx)].auto_size = True

    # Apply styles to the header
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Define column formats
    percentage_format = NamedStyle(name="percent_style", number_format="0%")          
            
    for col in ['G', 'I', 'K', 'M', 'R',]:
        for row in range(2, sheet.max_row + 1):
            cell = sheet[f'{col}{row}']
            cell.style = percentage_format
    

    workbook.save(buffer)

    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pfp.xlsx"'

    return response


@login_required
@permission_required('operations.access_operations', raise_exception=True)
def agent_productivity_create_pdf_file(request):
    files_instance = AgentProductivity.objects.last()

    if not files_instance:
        return JsonResponse({}, content_type="application/json", status=204)

    merged_with_calls = agent_productivity_process_data(files_instance)

    final_data_dict = merged_with_calls.to_dict(orient='records')

    percentage_columns = ['In Call (%)', 'Ready (%)', 'Wrapup (%)', 'Not Ready (%)', 'Conversion Rate',]

    context = {
        'data': final_data_dict,
        'columns': merged_with_calls.columns,
        'percentage_columns': percentage_columns,
    }

    html_string = render_to_string('operations/agent_productivity_pdf.html', context)

    pdf_file = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-DIsposition'] = 'inline; filename="output.pdf"'

    return response


@login_required
@permission_required('operations.access_operations', raise_exception=True)
def collector_goal_summary(request):
    yesterday = datetime.now() - timedelta(days=1)
    formatted_yesteday = yesterday.strftime('%Y-%m-%d') 
    return render(request, 'operations/collector_goal_summary.html', {'selected_date': formatted_yesteday})


@login_required
@permission_required('operations.access_operations', raise_exception=True)
def collector_goal_summary_get_table_data(request):
    todays_date = datetime.now(timezone.utc) - timedelta(days=1)

    if request.method == 'POST':
        cgs_post_data = json.loads(request.body)
        recieved_date = cgs_post_data['dateInput']
        if recieved_date:  # Us the selected date if one was sent in post request
            todays_date = datetime.strptime(recieved_date, '%Y-%m-%d')
            todays_date = todays_date.astimezone(timezone.utc)
    
    yesterdays_date = todays_date - timedelta(days=1)
    yesterdays_date = yesterdays_date.astimezone(timezone.utc)

    first_day = todays_date.replace(day=1)
    next_month = first_day.replace(month=first_day.month % 12 + 1, day=1)
    last_day = next_month - timedelta(days=1)

    startDate = first_day.strftime('%Y-%m-%d')
    endDate = last_day.strftime('%Y-%m-%d')

    # RETREIVE DATA #############################################################################################
    employee_data, ci_bigbank_postdates_data, mab_tum_postdates_data, posted_data = collector_goal_summary_retrieve_data(startDate, endDate)

    # TODO: CHECK IF IT"S BEGINIG OF MONTH< AND RETREIVE LAST MONTHS DATA IF NEEDED
    if todays_date.day == 1:
        first_day = yesterdays_date.replace(day=1)
        next_month = first_day.replace(month=first_day.month % 12 + 1, day=1)
        last_day = next_month - timedelta(days=1)

        startDate = first_day.strftime('%Y-%m-%d')
        endDate = last_day.strftime('%Y-%m-%d')

        _, yesterdays_ci_bigbank_postdates_data, yesterdays_mab_tum_postdates_data, yesterdays_posted_data = \
            collector_goal_summary_retrieve_data(startDate, endDate)
        _, _, yesterdays_inhouse_fees = collector_goal_summary_caluclate_collected_fee(yesterdays_date, yesterdays_ci_bigbank_postdates_data, 
                                                                                       yesterdays_mab_tum_postdates_data, yesterdays_posted_data)
    else:
        _, _, yesterdays_inhouse_fees = collector_goal_summary_caluclate_collected_fee(yesterdays_date, ci_bigbank_postdates_data, mab_tum_postdates_data, posted_data)

    
    posted_fees, pdc_fees, inhouse_fees = collector_goal_summary_caluclate_collected_fee(todays_date, ci_bigbank_postdates_data, mab_tum_postdates_data, posted_data)
    _, _, yesterdays_inhouse_fees = collector_goal_summary_caluclate_collected_fee(yesterdays_date, ci_bigbank_postdates_data, mab_tum_postdates_data, posted_data)



    # Subtract yesterdays inhouse fees from today's inhouse fees to get total move
    inhouse_combined = pd.merge(yesterdays_inhouse_fees, inhouse_fees, on='AssociateID', how='outer', suffixes=('_left', '_right')).fillna(0)
    inhouse_combined['Move'] = inhouse_combined['Inhouse Fees_left'] - inhouse_combined['Inhouse Fees_right']
    move = inhouse_combined.drop(['Inhouse Fees_left', 'Inhouse Fees_right'], axis=1).fillna(0)
    # TODO: MOVE BREAKS ON FIRST DAY OF MONTH BECUASE THE DATA RECIEVED IS ONLY FOR THE CURRENT MONTH

    # Get the LegacyDeskID and GoalAmount from posted_data DeskID    
    ci_bigbank_postdates_data_desk_data = ci_bigbank_postdates_data[['DeskID', 'LegacyDeskID', 'GoalAmount']]
    mab_tum_postdates_data_desk_data = mab_tum_postdates_data[['DeskID', 'LegacyDeskID', 'GoalAmount']]
    posted_data_desk_data = posted_data[['DeskID', 'LegacyDeskID', 'GoalAmount']]
    desk_data = pd.concat([ci_bigbank_postdates_data_desk_data, mab_tum_postdates_data_desk_data, posted_data_desk_data])
    desk_data = desk_data.drop_duplicates(subset=['DeskID'])
    desk_data = pd.merge(desk_data, employee_data[['DeskID', 'EmployeeID', 'AssociateID']], on='DeskID', how='left')
    # desk_data = desk_data[(desk_data['GoalAmount'] != 0) | (desk_data['EmployeeID'] != 0)]

    



    # Add all the fee data into one dataframe
    final_data = pd.merge(posted_fees, pdc_fees, on='AssociateID', how='outer')
    final_data = pd.merge(final_data, inhouse_fees, on='AssociateID', how='outer')
    final_data = pd.merge(final_data, move, on='AssociateID', how='outer')
    # final_data = pd.merge(final_data, employee_data[['EmployeeID', 'AssociateID']], on='AssociateID', how='left')
    # Sum up fees for each desk by employee id
    final_data = final_data.groupby('AssociateID').agg({'Posted Fees': 'sum', 'PDC Fees': 'sum', 'Inhouse Fees': 'sum', 'Move': 'sum'}).reset_index()
    # Add values from employees data
    final_data = pd.merge(final_data, employee_data[['ZID', 'FirstName', 'LastName', 'Department', 'EmployeeID', 'AssociateID']], on='AssociateID', how='right')
    # final_data = pd.merge(final_data, desk_data[['LegacyDeskID', 'GoalAmount', 'DeskID']], on='DeskID', how='left')
    final_data = pd.merge(final_data, desk_data[['LegacyDeskID', 'GoalAmount', 'AssociateID']], on='AssociateID', how='left')  # Add desk data to the employee data
    # Drop duplicates. Assumes there is always a complete row of data if GoalAmount is not zero, unless it's a Redial employee. Assumes there arern't duplicate rows for Redial.
    final_data = final_data[(final_data['GoalAmount'] != 0) | (final_data['Department'] == 'REDIAL')]  
    final_data = final_data[final_data['Department'] != 'Naclientsh']




    final_data = final_data.drop_duplicates().reset_index(drop=True)  # the employee data includes all unique desk ids, so their will be duplicate EmployeeIDs
    final_data.fillna(0, inplace=True)

    # Add rank column
    final_data['Rank'] = final_data['Inhouse Fees'].rank(ascending=False, method='min')
    final_data = final_data.sort_values(by='Rank')

    # Reorder columns and Rename where needed 
    new_column_order = ['Rank', 'ZID', 'FirstName', 'LastName', 'Department', 'Posted Fees', 'PDC Fees', 'Inhouse Fees', 'Move', 'GoalAmount', 'LegacyDeskID', 'EmployeeID', 'AssociateID']
    final_data = final_data[new_column_order]
    new_column_names = {
        'EmployeeID': 'Emp ID',
        'AssociateID': 'Assoc ID',
        'LegacyDeskID': 'Desk',
        'GoalAmount': 'Goal',
        'FirstName': 'First Name', 
        'LastName': 'Last Name',
    }
    final_data = final_data.rename(columns=new_column_names)

    collector_goal_summary_dict = final_data.to_dict(orient='records')
    return JsonResponse(collector_goal_summary_dict, safe=False)


def collector_goal_summary_retrieve_data(startDate, endDate):
    api_url = f'http://host.docker.internal:3000/api/detailed_employees'

    try:
        response = requests.get(api_url, headers={ 'x-api-key': os.getenv('API_KEY') })
        response.raise_for_status()

        data = response.json()

    except requests.exceptions.RequestException as e:
        print(e)
        pass
        # TODO: THIS PROBABLY WON"T WORK HERE
        #messages.error(request, f"An error occured using the KPI API: {str(e)}.")

    employee_data = pd.DataFrame(data)
 

    api_url = f'http://host.docker.internal:3000/api/ci_bigbank_postdates?startDate={startDate}&endDate={endDate}'

    try:
        response = requests.get(api_url, headers={ 'x-api-key': os.getenv('API_KEY') })
        response.raise_for_status()

        data = response.json()

    except requests.exceptions.RequestException as e:
        pass
        # TODO: THIS PROBABLY WON"T WORK HERE
        #messages.error(request, f"An error occured using the KPI API: {str(e)}.")

    ci_bigbank_postdates_data = pd.DataFrame(data)
    ci_bigbank_postdates_data['DateCreated'] = pd.to_datetime(ci_bigbank_postdates_data['DateCreated'], utc=True)
    ci_bigbank_postdates_data['DateDeleted'] = pd.to_datetime(ci_bigbank_postdates_data['DateDeleted'], utc=True)
    ci_bigbank_postdates_data['DepositDate'] = pd.to_datetime(ci_bigbank_postdates_data['DepositDate'], utc=True)
    print(ci_bigbank_postdates_data.columns)


    api_url = f'http://host.docker.internal:3000/api/mab_tum_postdates?startDate={startDate}&endDate={endDate}'

    try:
        response = requests.get(api_url, headers={ 'x-api-key': os.getenv('API_KEY') })
        response.raise_for_status()

        data = response.json()

    except requests.exceptions.RequestException as e:
        pass
        # TODO: THIS PROBABLY WON"T WORK HERE
        #messages.error(request, f"An error occured using the KPI API: {str(e)}.")

    mab_tum_postdates_data = pd.DataFrame(data)
    mab_tum_postdates_data['DateCreated'] = pd.to_datetime(mab_tum_postdates_data['DateCreated'], utc=True)
    mab_tum_postdates_data['DateDeleted'] = pd.to_datetime(mab_tum_postdates_data['DateDeleted'], utc=True)
    mab_tum_postdates_data['DepositDate'] = pd.to_datetime(mab_tum_postdates_data['DepositDate'], utc=True)

    test = pd.merge(mab_tum_postdates_data, employee_data[['EmployeeID', 'AssociateID', 'CUBSInitials']], left_on='LegacyDeskID', right_on='CUBSInitials', how='left')
    test = test.drop('CUBSInitials', axis=1)
    print(test.columns)
    

    api_url = f'http://host.docker.internal:3000/api/detailed_posted?startDate={startDate}&endDate={endDate}'

    try:
        response = requests.get(api_url, headers={ 'x-api-key': os.getenv('API_KEY') })
        response.raise_for_status()

        data = response.json()

    except requests.exceptions.RequestException as e:
        pass
        # TODO: THIS PROBABLY WON"T WORK HERE
        #messages.error(request, f"An error occured using the KPI API: {str(e)}.")

    # Calculate Posted Fees
    posted_data = pd.DataFrame(data)
    posted_data['PostingDateTime'] = pd.to_datetime(posted_data['PostingDateTime'], utc=True)

    return employee_data, ci_bigbank_postdates_data, mab_tum_postdates_data, posted_data


def collector_goal_summary_caluclate_collected_fee(target_date, ci_bigbank_postdates_data, mab_tum_postdates_data, posted_data):
    if not posted_data.empty:
        posted_fees = posted_data[posted_data['PostingDateTime'] < target_date].groupby('AssociateID')['CommissionAmount'].sum().reset_index().fillna(0) \
            .rename(columns={'CommissionAmount': 'Posted Fees'})
    else:
        posted_fees = posted_data.fillna(0)

    combined_pdc_fees = pd.concat([ci_bigbank_postdates_data, mab_tum_postdates_data], ignore_index=True)
    left_combined_pdc_fees = combined_pdc_fees[(combined_pdc_fees['DepositDate'] >= target_date) & (combined_pdc_fees['DateCreated'] < target_date)] \
        .groupby('AssociateID')['ProjectedCommissionAmount'].sum().reset_index().fillna(0)    
    right_combined_pdc_fees = combined_pdc_fees[(combined_pdc_fees['DepositDate'] >= target_date) & (combined_pdc_fees['DateCreated'] < target_date) & 
                                                  (combined_pdc_fees['DateDeleted'] < target_date)] \
        .groupby('AssociateID')['ProjectedCommissionAmount'].sum().reset_index().fillna(0)

    # Subtract left from right pdc fees
    merged_pdc_fees = pd.merge(left_combined_pdc_fees, right_combined_pdc_fees, on='AssociateID', how='outer', suffixes=('_left', '_right')).fillna(0)
    merged_pdc_fees['PDC Fees'] = merged_pdc_fees['ProjectedCommissionAmount_left'] - merged_pdc_fees['ProjectedCommissionAmount_right']
    pdc_fees = merged_pdc_fees.drop(['ProjectedCommissionAmount_left', 'ProjectedCommissionAmount_right'], axis=1).fillna(0)

    # Calculate in house fees
    if not posted_data.empty:
        merged_inhouse_fees = pd.merge(posted_fees, pdc_fees, on='AssociateID', how='outer')
        merged_inhouse_fees.fillna(0, inplace=True)
        merged_inhouse_fees['Inhouse Fees'] = merged_inhouse_fees['Posted Fees'] + merged_inhouse_fees['PDC Fees']
        inhouse_fees = merged_inhouse_fees.drop(['Posted Fees', 'PDC Fees'], axis=1).fillna(0)
    else:
        inhouse_fees = pdc_fees.rename(columns={'PDC Fees': 'Inhouse Fees'}).fillna(0)

    return posted_fees, pdc_fees, inhouse_fees
