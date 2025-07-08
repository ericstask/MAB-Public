import math
import requests
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from datetime import datetime, timedelta
from io import BytesIO
from .models import PayForPerformanceEmployee

def create_pfp_spreadsheet(dataframe):
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        dataframe.to_excel(writer, index=False)

    # Seek the beginning of the buffer so it can be read from the start
    buffer.seek(0)

    wb = load_workbook(buffer)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    blue_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
    thin_boarder = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_alignment = Alignment(horizontal='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center')

    sheet = wb.active   

    # Double height of header row
    sheet.row_dimensions[1].height = 34

    # Make the header column text green
    for cell in sheet[1]:
        cell.font = Font(color='006100')
        cell.alignment = header_alignment

    # Apply accounting number format to currecy fields
    accounting_style = NamedStyle(name='accounting_style', number_format=r'_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"_);_(@_)')
    for column in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',]:
        for cell in sheet[column][1:]:
            cell.style = accounting_style

    # Color rolling average cells
    for cell in sheet['L']:
        cell.fill = yellow_fill
    for cell in sheet['M']:
        cell.fill = blue_fill

    # Format Date column to MM/DD/YYYY
    for cell in sheet['D'][1:]:
        cell.number_format = 'MM/DD/YYYY'

    # Center align specific columns
    for col in [2, 3, 4, 5]:
        for cell in sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=col, max_col=col):
            for c in cell:
                c.alignment = center_alignment

    # Set with of columns to best fit
    column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(sheet.max_column))
    for column_letter in column_letters:
        sheet.column_dimensions[column_letter].bestFit = True

    # Add borders to spreadsheet
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_boarder

    # Set the AutoFilter for the range of cells
    max_row = sheet.max_row
    max_column = sheet.max_column
    start_cell = 'A1'
    end_cell = get_column_letter(max_column) + str(max_row)
    sheet.auto_filter.ref = f'{start_cell}:{end_cell}'

    wb.save(buffer)

    # reset pointer again
    buffer.seek(0)

    return buffer


def create_pfp_dataframe(employee_data, transaction_data):
    MONTH_NAMES = {
        1: 'January',
        2: 'February',
        3: 'March',
        4: 'April',
        5: 'May',
        6: 'June',
        7: 'July',
        8: 'August',
        9: 'September',
        10: 'October',
        11: 'November',
        12: 'December',
    }

    current_date = datetime.now()

    # Convert hire date column to datetime type
    employee_data['Most Recent Hire Date'] = pd.to_datetime(employee_data['Most Recent Hire Date'])
    
    # Combine first name and last name into one column
    employee_data['Full Name'] = employee_data['Employee Last Name and Suffix'] + ', ' + employee_data['Employee First Name']
    employee_data['Tenure'] = (current_date.year - employee_data['Most Recent Hire Date'].dt.year) * 12 + (current_date.month - employee_data['Most Recent Hire Date'].dt.month)

    # Extract Month and Year from posting date
    transaction_data['PostingDateTime'] = pd.to_datetime(transaction_data['PostingDateTime'])

    transaction_data['Month'] = transaction_data['PostingDateTime'].dt.month
    transaction_data['Year'] = transaction_data['PostingDateTime'].dt.year

    # Combine Month and Year into one column for sorting
    transaction_data['Month_Year'] = transaction_data['Month'].astype(str) + '_' + transaction_data['Year'].astype(str)

    # Group by EmployeeID and Month, and calculate the total number of transactions for each group
    transactions_summary = transaction_data.groupby(['EmployeeID', 'Month_Year'])['CommissionAmount'].sum().reset_index()

    # Pivot the data to have EmployeeID as index, Month as colums, and Comission as values
    commission_summary_pivot = transactions_summary.pivot(index='EmployeeID', columns='Month_Year', values='CommissionAmount').fillna(0)

    # Sort columns based on Month_Year
    commission_summary_pivot = commission_summary_pivot.reindex(columns=sorted(commission_summary_pivot.columns, key=lambda x: tuple(map(int, x.split('_')[::-1]))))

    # Rename Month columns
    commission_summary_pivot.columns = [MONTH_NAMES[int(col.split('_')[0])] for col in commission_summary_pivot.columns]

    #print(commission_summary_pivot[commission_summary_pivot['EmployeeID'] == '14695'])

    # Calculate rolling averages
    commission_summary_pivot['90 Day Rolling Average'] = commission_summary_pivot.iloc[:,-3:].mean(axis=1)
    commission_summary_pivot['Prior 90 Day Rolling Average'] = commission_summary_pivot.iloc[:,-7:-4:].mean(axis=1)

    # Get the months in the table before we add the EmployeeID back as a column
    additional_columns_to_include = list(commission_summary_pivot.columns)

    # Reset the index to make the EmployeeID a column instead of an index
    commission_summary_pivot.reset_index(inplace=True)

    # Change the datatype of employee id to int so it matches the data type in the employee data
    commission_summary_pivot['EmployeeID'] = commission_summary_pivot['EmployeeID'].astype(int)

    # Merge the commission amounts with the employee dataframe
    employee_data = pd.merge(employee_data, commission_summary_pivot, left_on='Employee ID', right_on='EmployeeID', how='left')

    # Remove first party employees
    employee_data = employee_data.dropna(subset=['Prior 90 Day Rolling Average'])

    # Change Most Recent Hire Date format from datetime to date
    employee_data['Most Recent Hire Date'] = employee_data['Most Recent Hire Date'].dt.date

    
    columns_to_include = ['Full Name', 'Employee ID', 'Supervisor', 'Most Recent Hire Date', 'Tenure',] + additional_columns_to_include + ['Pay Rate 1']
    final_data = employee_data[columns_to_include]
    
    final_data = final_data.sort_values('Full Name')

    # Convert Employee ID from decimal number to integer
    final_data["Employee ID"] = final_data["Employee ID"].astype(int)

    new_column_names = {
        'Full Name': 'Collector Name',
        'Employee ID': 'EE ID',
        'Supervisor': 'Manager',
        'Most Recent Hire Date': 'Start Date',
        'Tenure': 'Tenure (Mo.)',
        'Pay Rate 1': 'Hourly Wage',
    }

    final_data = final_data.rename(columns=new_column_names)

    return final_data


def pay_for_performance_query_transactions(df, startDate, endDate):
    # Drop rows that lack an employee id
    df = df.dropna(subset=['Employee ID'])

    # Remove employees from Paychex report that are not wanted
    removed_employees = list(PayForPerformanceEmployee.objects.filter(status='removed', is_removed=False).values_list('id', flat=True))
    df = df[~df['Employee ID'].isin(removed_employees)]

    # Add new employees that aren't on Paychex
    added_employees = list(PayForPerformanceEmployee.objects.filter(status='added', is_removed=False).values('id', 'last_name_and_suffix', 'first_name', 
                                                                                                                'most_recent_hire_date', 'supervisor', 'pay_rate'))
    if added_employees:
        new_column_names = {
            'id': 'Employee ID',
            'last_name_and_suffix': 'Employee Last Name and Suffix',
            'first_name': 'Employee First Name',
            'most_recent_hire_date': 'Most Recent Hire Date',
            'supervisor': 'Supervisor',
            'pay_rate': 'Pay Rate 1',
        }
        added_employees_df = pd.DataFrame(added_employees)
        added_employees_df['pay_rate'] = pd.to_numeric(added_employees_df['pay_rate'], errors='coerce') # make sure the pay rate is stored as a number
        added_employees_df = added_employees_df.rename(columns=new_column_names)

        # Combine added employees with Paychex data
        combined_df = pd.concat([df, added_employees_df], axis=0, ignore_index=True, sort=False)

        # Get employee ids to send in api request
        data_dict = combined_df.to_dict(orient='list')
    else: 
        data_dict = df.to_dict(orient='list')
    
    employee_ids = [int(num) for num in data_dict["Employee ID"] if not math.isnan(num)]

    # Get data from api
    api_url = f'http://host.docker.internal:3000/api/pfp'

    data = {
        'employeeIds': employee_ids,
        'startDate': startDate,
        'endDate': endDate,
    }

    try:
        response = requests.post(api_url, json=data)
        response.raise_for_status()

        data = response.json()

    except requests.exceptions.RequestException as e:
        pass
        # TODO: THIS PROBABLY WON"T WORK HERE
        #messages.error(request, f"An error occured using the KPI API: {str(e)}.")


    transaction_data = pd.DataFrame(data)

    if added_employees:
        return combined_df, transaction_data
    else:
        return df, transaction_data


def agent_productivity_process_data(files_instance):
    productivity_df = pd.read_excel(files_instance.productivity_uploaded_file, skiprows=2)
    calls_df = pd.read_excel(files_instance.calls_uploaded_file, skiprows=2)

    payment_codes = ['PAY1', 'PAY2', 'PAY3', 'PAY4', 'PAY5',]
    non_rpc_codes_to_remove = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P',
                            'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',]
    
    # Remove whitespace from data
    calls_df = calls_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

    # Remove calls that are not RPCs
    calls_df = calls_df[~calls_df['Result Code'].isin(non_rpc_codes_to_remove)]

    # Count total number of calls for each agent
    call_counts = calls_df.groupby('Agent').size().reset_index(name='Total RPC')
    # Count the number of rows for each agent that resulted in a payment
    payment_counts = calls_df[calls_df['Result Code'].isin(payment_codes)].groupby('Agent').size().reset_index(name='Payment Count')
    # Merge Counts to get percentage
    merged_calls_df = pd.merge(call_counts, payment_counts, on='Agent', how='left')
    # Fill NaN with a 0
    merged_calls_df['Payment Count'] = merged_calls_df['Payment Count'].fillna(0)
    merged_calls_df['Conversion Rate'] = (merged_calls_df['Payment Count'] / merged_calls_df['Total RPC'])


    # Remove underscores from Agent names
    productivity_df['Agent_cleaned'] = productivity_df['Agent'].apply(clean_agent_name_in_calls)
    merged_with_calls = pd.merge(productivity_df, merged_calls_df, left_on='Agent_cleaned', right_on='Agent', how='left')

    columns_to_drop = ['Agent_cleaned', 'Agent_y', 'Transfers', 'PD', 'Total WPC', 'Total RPC_y']
    merged_with_calls = merged_with_calls.drop(columns=columns_to_drop)

    merged_with_calls = merged_with_calls.rename(columns={'Agent_x': 'Agent', 'Total RPC_x': 'Total RPC'})

    merged_with_calls = merged_with_calls.fillna(0)

    return merged_with_calls


def clean_agent_name_in_calls(agent_name):
    try:
        first_name, last_name = agent_name.split('_')[0].split()
        return (first_name[0] + last_name).upper()
    except ValueError:
        return None
    